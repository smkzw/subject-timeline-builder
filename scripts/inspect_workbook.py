#!/usr/bin/env python3
import argparse
import json
from collections import Counter
from pathlib import Path

import openpyxl


def suffixed(headers):
    seen = Counter()
    out = []
    for h in headers:
        name = "" if h is None else str(h)
        seen[name] += 1
        out.append(name if seen[name] == 1 else f"{name}__{seen[name]}")
    return out


def main():
    ap = argparse.ArgumentParser(description="Inspect workbook sheets, headers, duplicate headers, and sample rows.")
    ap.add_argument("workbook")
    ap.add_argument("--sample-rows", type=int, default=2)
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info = {"file": str(path), "sheets": []}
    for ws in wb.worksheets:
        raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = suffixed(raw_headers)
        counts = Counter("" if h is None else str(h) for h in raw_headers)
        duplicates = [h for h, n in counts.items() if h and n > 1]
        samples = []
        for row in ws.iter_rows(min_row=2, max_row=1 + args.sample_rows, values_only=True):
            samples.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        info["sheets"].append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": headers,
            "duplicate_headers": duplicates,
            "samples": samples,
        })
    wb.close()

    if args.json:
        print(json.dumps(info, ensure_ascii=False, indent=2, default=str))
        return

    print(f"Workbook: {info['file']}")
    for sheet in info["sheets"]:
        print(f"\n[{sheet['name']}] rows={sheet['max_row']} cols={sheet['max_column']}")
        if sheet["duplicate_headers"]:
            print("duplicate headers:", ", ".join(sheet["duplicate_headers"]))
        print("headers:")
        for idx, header in enumerate(sheet["headers"], 1):
            print(f"  {idx}. {header}")
        if sheet["samples"]:
            print("sample:")
            for sample in sheet["samples"]:
                print(" ", json.dumps(sample, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
