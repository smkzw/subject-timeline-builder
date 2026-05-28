#!/usr/bin/env python3
import argparse
import calendar
import datetime as dt
import html
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


def esc(value):
    return html.escape(str(value)) if value not in (None, "") else ""


def suffixed(headers):
    seen = Counter()
    out = []
    for h in headers:
        name = "" if h is None else str(h)
        seen[name] += 1
        out.append(name if seen[name] == 1 else f"{name}__{seen[name]}")
    return out


def parse_partial_date(value, bound):
    if value in (None, "", "None", "NA", "N/A", "未在现有文件中确认"):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip().replace("\t", "").replace("\n", "").replace("/", "-").replace(".", "-")
    if not text or "ongoing" in text.lower() or "持续" == text:
        return None
    parts = text.split("-")
    try:
        year = int(parts[0])
        month = 1
        day = 1
        if len(parts) >= 2:
            month = 1 if parts[1].lower() == "uk" else int(parts[1])
        if len(parts) >= 3:
            if parts[2].lower() == "uk":
                day = 1 if bound == "start" else calendar.monthrange(year, month)[1]
            else:
                day = int(parts[2])
        elif bound == "end":
            day = calendar.monthrange(year, month)[1] if len(parts) >= 2 else 31
            month = month if len(parts) >= 2 else 12
        return dt.date(year, month, day)
    except Exception:
        return None


def load_sheet(path, sheet_cfg):
    if not sheet_cfg:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_cfg["name"]]
    headers = suffixed(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rec = {headers[i]: ("" if values[i] is None else str(values[i])) for i in range(min(len(headers), len(values)))}
        rows.append(rec)
    wb.close()
    return rows


def col(cfg, key):
    return cfg.get(key, "")


def val(row, cfg, key, default=""):
    name = col(cfg, key)
    return row.get(name, default) if name else default


def subject_key(row, cfg):
    return val(row, cfg, "subject").strip()


def center_value(row, cfg):
    return val(row, cfg, "center").strip() or "未识别中心"


def truthy(value):
    return str(value or "").strip() in ("是", "Yes", "Y", "1", "TRUE", "True")


def is_screen_failed_status(value):
    text = str(value or "")
    return any(x in text for x in ["筛选失败", "筛败", "未入组", "未随机", "Screen Failure"])


def subject_status(row, cfg):
    for key in ("status", "subject_status", "screen_status", "random_status"):
        if col(cfg, key):
            text = val(row, cfg, key).strip()
            if text:
                return text
    return ""


def normalize_visit(visit, visit_order):
    text = str(visit or "")
    for item in visit_order:
        if item.get("match", "") in text:
            return item.get("label", text), int(item.get("order", 999))
    return text, 999


def sorted_visit_dates(subject, visits, visits_cfg, visit_order):
    rows = []
    for row in visits.get(subject, []):
        if row.get("__is_usv") and not visits_cfg.get("include_usv", False):
            continue
        occurred_col = col(visits_cfg, "occurred")
        if occurred_col and row.get(occurred_col) not in ("", "是", "Yes", "Y", "1", "TRUE", "True"):
            continue
        d = parse_partial_date(val(row, visits_cfg, "date"), "start")
        if d:
            label, order = normalize_visit(val(row, visits_cfg, "visit"), visit_order)
            rows.append((d, label, order, row))
    return sorted(rows, key=lambda x: (x[0], x[2]))


def window_from_visits(subject, visits, visits_cfg, window_cfg, visit_order):
    dated = sorted_visit_dates(subject, visits, visits_cfg, visit_order)
    if not dated:
        return None, None
    first, last = dated[0][0], dated[-1][0]
    wtype = (window_cfg or {}).get("type", "first_visit_to_last_visit")
    if wtype == "first_visit_to_last_visit":
        return first, last
    if wtype in ("first_dose_to_last_visit", "first_dose_to_treatment_end"):
        start_match = (window_cfg or {}).get("start_visit_match", "基线")
        start = next((d for d, label, _, row in dated if start_match in label or start_match in val(row, visits_cfg, "visit")), first)
        if wtype == "first_dose_to_last_visit":
            return start, last
        end_match = (window_cfg or {}).get("end_visit_match", "治疗结束")
        end = next((d for d, label, _, row in reversed(dated) if end_match in label or end_match in val(row, visits_cfg, "visit")), None)
        if not end:
            end = next((d for d, label, _, row in reversed(dated) if "24" in label or "24" in val(row, visits_cfg, "visit")), last)
        return start, end
    return first, last


def overlaps_window(row, cfg, start_col, end_col, ongoing_col, window_start, window_end):
    if not window_start or not window_end:
        return False
    start = parse_partial_date(row.get(start_col, ""), "start")
    ongoing = row.get(ongoing_col, "") in ("是", "Yes", "Y", "1", "TRUE", "True")
    end = window_end if ongoing else (parse_partial_date(row.get(end_col, ""), "end") or start)
    if not start and not end:
        return False
    if not start:
        start = end
    if not end:
        end = window_end
    return start <= window_end and end >= window_start


def is_research_drug(name, patterns):
    return any(p and p in name for p in patterns)


def li(text, title=""):
    return f'<li title="{esc(title)}">{text}</li>'


def render_list(items, empty):
    return "<ul>" + ("".join(items) if items else f"<li>{esc(empty)}</li>") + "</ul>"


def cm_reason(row, cfg):
    reason = val(row, cfg, "reason")
    reason_mh = val(row, cfg, "reason_mh")
    reason_ae = val(row, cfg, "reason_ae")
    other = val(row, cfg, "reason_other")
    note = val(row, cfg, "note")
    details = [row.get(c, "") for c in cfg.get("reason_detail_columns", []) if row.get(c, "")]
    if reason_ae:
        return f"AE: {reason_ae}"
    if reason_mh:
        return f"病史: {reason_mh}"
    if details:
        return f"{reason}: " + "；".join(details)
    if any(x in reason for x in ["不良事件", "既往", "现病史", "病史"]) and note:
        return f"{reason}: {note}"
    return other or reason


def cm_reason_needs_confirmation(row, cfg):
    reason = val(row, cfg, "reason")
    linked_reason = "不良事件" in reason or ("其他" in reason and any(x in reason for x in ["既往", "现病史", "病史"]))
    if not linked_reason:
        return False
    reason_mh = val(row, cfg, "reason_mh")
    reason_ae = val(row, cfg, "reason_ae")
    note = val(row, cfg, "note")
    details = [row.get(c, "") for c in cfg.get("reason_detail_columns", []) if row.get(c, "")]
    return not any([reason_mh, reason_ae, note, details])


def load_group_info(path, group_cfg, skip_subjects):
    info = {}
    if not path or not group_cfg:
        return info
    for row in load_sheet(path, group_cfg):
        sid = subject_key(row, group_cfg)
        if not sid or sid in skip_subjects:
            continue
        group = val(row, group_cfg, "group")
        status = subject_status(row, group_cfg)
        random_no = val(row, group_cfg, "random_no")
        label = "筛败" if is_screen_failed_status(status) else (group or ("已随机" if random_no else ""))
        info[sid] = {
            "group": group,
            "status": status,
            "random_no": random_no,
            "label": label or status,
            "is_randomized": bool(group or random_no) and not is_screen_failed_status(status),
            "is_screen_failed": is_screen_failed_status(status),
        }
    return info


def display_subject(sid, group_info, fallback_status=""):
    info = group_info.get(sid, {})
    label = info.get("label") or ("筛败" if is_screen_failed_status(fallback_status) else "")
    return f"{sid} | {label}" if label else sid


def render_svg(subject, visit_rows, ae_rows, mh_rows, cm_rows, cfg, window_start, window_end, visit_order):
    if not window_start or not window_end:
        return ""
    width, left, right = 1500, 180, 260
    def x_pos(d):
        if not d:
            return left
        d = max(min(d, window_end), window_start)
        span = max((window_end - window_start).days, 1)
        return left + ((d - window_start).days / span) * (width - left - right)

    visits = []
    for row in visit_rows:
        d = parse_partial_date(val(row, cfg["visits"], "date"), "start")
        label, order = normalize_visit(val(row, cfg["visits"], "visit"), visit_order)
        if d:
            visits.append((order, label, val(row, cfg["visits"], "date"), d))
    visits.sort(key=lambda x: (x[0], x[3]))

    ae_y = 160
    mh_y = ae_y + max(len(ae_rows), 1) * 20 + 34
    cm_y = mh_y + max(len(mh_rows), 1) * 18 + 40
    height = cm_y + max(len(cm_rows), 1) * 18 + 42
    parts = [
        f'<line x1="{left}" y1="86" x2="{width-right}" y2="86" stroke="#d6c29b" stroke-width="4"/>',
        f'<text x="{left}" y="28" font-size="12" fill="#666">{esc(window_start)}</text>',
        f'<text x="{width-right}" y="28" text-anchor="end" font-size="12" fill="#666">{esc(window_end)}</text>',
        f'<text x="24" y="{ae_y-14}" font-size="16" font-weight="700" fill="#444">AE / 安全性事件</text>',
        f'<text x="24" y="{mh_y-8}" font-size="16" font-weight="700" fill="#444">病史</text>',
        f'<text x="24" y="{cm_y-8}" font-size="16" font-weight="700" fill="#444">合并用药</text>',
    ]
    for _, label, raw, d in visits:
        x = x_pos(d)
        parts += [
            f'<line x1="{x:.1f}" y1="76" x2="{x:.1f}" y2="{height-36}" stroke="#efe7d6" stroke-width="1"/>',
            f'<line x1="{x:.1f}" y1="76" x2="{x:.1f}" y2="96" stroke="#8f7c58" stroke-width="2"/>',
            f'<text x="{x:.1f}" y="110" text-anchor="middle" font-size="11" fill="#444">{esc(label)}</text>',
            f'<text x="{x:.1f}" y="126" text-anchor="middle" font-size="10" fill="#777">{esc(raw)}</text>',
        ]

    ae_cfg, mh_cfg, cm_cfg = cfg.get("ae", {}), cfg.get("mh", {}), cfg.get("cm", {})
    for idx, row in enumerate(ae_rows, 1):
        sd = parse_partial_date(val(row, ae_cfg, "start"), "start")
        ed = window_end if val(row, ae_cfg, "ongoing") in ("是", "Yes", "Y") else parse_partial_date(val(row, ae_cfg, "end"), "end") or sd
        x1, x2, y = x_pos(sd), x_pos(ed), ae_y + (idx - 1) * 20
        title = f"{val(row, ae_cfg, 'term')} | {val(row, ae_cfg, 'start')} ~ {val(row, ae_cfg, 'end')} | {val(row, ae_cfg, 'grade')} | SAE {val(row, ae_cfg, 'sae')}"
        parts.append(f'<text x="170" y="{y+14}" text-anchor="end" font-size="11" fill="#444">AE#{idx}</text>')
        parts.append(f'<rect x="{x1:.1f}" y="{y}" width="{max(4, x2-x1):.1f}" height="18" rx="3" fill="#b9d6ec" stroke="#666"><title>{esc(title)}</title></rect>')
    for idx, row in enumerate(mh_rows, 1):
        sd = parse_partial_date(val(row, mh_cfg, "start"), "start")
        ed = window_end if val(row, mh_cfg, "ongoing") in ("是", "Yes", "Y") else parse_partial_date(val(row, mh_cfg, "end"), "end") or sd
        x1, x2, y = x_pos(sd), x_pos(ed), mh_y + (idx - 1) * 18
        parts.append(f'<text x="170" y="{y+11}" text-anchor="end" font-size="11" fill="#444">MH#{idx}</text>')
        parts.append(f'<rect x="{x1:.1f}" y="{y}" width="{max(4, x2-x1):.1f}" height="15" rx="2" fill="#d9d0ff" stroke="#6b58a6"><title>{esc(val(row, mh_cfg, "term"))}</title></rect>')
    for idx, row in enumerate(cm_rows, 1):
        sd = parse_partial_date(val(row, cm_cfg, "start"), "start")
        ed = window_end if val(row, cm_cfg, "ongoing") in ("是", "Yes", "Y") else parse_partial_date(val(row, cm_cfg, "end"), "end") or sd
        x1, x2, y = x_pos(sd), x_pos(ed), cm_y + (idx - 1) * 18
        parts.append(f'<text x="170" y="{y+11}" text-anchor="end" font-size="11" fill="#444">CM#{idx}</text>')
        parts.append(f'<rect x="{x1:.1f}" y="{y}" width="{max(4, x2-x1):.1f}" height="15" rx="2" fill="#7fc97f" stroke="#4f7b46"><title>{esc(val(row, cm_cfg, "drug_name"))}</title></rect>')
    return f'<svg viewBox="0 0 {width} {height}" class="svgline">{"".join(parts)}</svg>'


def main():
    ap = argparse.ArgumentParser(description="Build Chinese interactive subject timeline HTML from mapped clinical listings.")
    ap.add_argument("--listing", required=True)
    ap.add_argument("--config", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--pk-listing")
    ap.add_argument("--finding-listing")
    ap.add_argument("--group-listing")
    args = ap.parse_args()

    config = json.loads(Path(args.config).read_text(encoding="utf-8"))
    sheets = config["sheets"]
    modules = config.get("modules", {})
    visit_order = config.get("visit_order", [])
    patterns = config.get("research_drug_name_patterns", [])
    skip_subjects = set(config.get("skip_subject_values", ["SUBJID", "Subject", "SUBJECT", "受试者"]))
    population = config.get("population", "all")
    sheets["visits"] = dict(sheets["visits"])
    sheets["visits"]["include_usv"] = config.get("include_usv", False)

    visits_rows = load_sheet(args.listing, sheets["visits"])
    visits = defaultdict(list)
    centers = {}
    statuses = {}
    for row in visits_rows:
        sid = subject_key(row, sheets["visits"])
        if sid and sid not in skip_subjects:
            visits[sid].append(row)
            centers[sid] = center_value(row, sheets["visits"])
            statuses.setdefault(sid, subject_status(row, sheets["visits"]))

    if config.get("include_usv", False) and sheets.get("usv"):
        usv_cfg = sheets["usv"]
        for row in load_sheet(args.listing, usv_cfg):
            sid = subject_key(row, usv_cfg)
            d = val(row, usv_cfg, "date")
            if sid and sid not in skip_subjects and d:
                row = dict(row)
                row["__is_usv"] = "1"
                row[col(sheets["visits"], "subject")] = sid
                row[col(sheets["visits"], "center")] = center_value(row, usv_cfg)
                row[col(sheets["visits"], "date")] = d
                row[col(sheets["visits"], "visit")] = usv_cfg.get("label", "USV")
                row["__usv_label"] = usv_cfg.get("label", "USV")
                visits[sid].append(row)
                centers.setdefault(sid, center_value(row, usv_cfg))
                statuses.setdefault(sid, subject_status(row, usv_cfg))

    group_info = load_group_info(args.group_listing or args.listing, sheets.get("group"), skip_subjects)
    if population == "randomized" and not group_info:
        raise SystemExit("population=randomized requires a group/randomization mapping or --group-listing.")

    grouped = {"ae": defaultdict(list), "mh": defaultdict(list), "cm": defaultdict(list), "pk": defaultdict(list), "findings": defaultdict(list)}
    if modules.get("ae") and sheets.get("ae"):
        for row in load_sheet(args.listing, sheets["ae"]):
            term = val(row, sheets["ae"], "term")
            sid = subject_key(row, sheets["ae"])
            if sid and sid not in skip_subjects and term:
                grouped["ae"][sid].append(row)
                centers.setdefault(sid, center_value(row, sheets["ae"]))
                statuses.setdefault(sid, subject_status(row, sheets["ae"]))
    if modules.get("mh") and sheets.get("mh"):
        for row in load_sheet(args.listing, sheets["mh"]):
            sid = subject_key(row, sheets["mh"])
            if sid and sid not in skip_subjects and val(row, sheets["mh"], "term"):
                grouped["mh"][sid].append(row)
                centers.setdefault(sid, center_value(row, sheets["mh"]))
                statuses.setdefault(sid, subject_status(row, sheets["mh"]))
    if modules.get("cm") and sheets.get("cm"):
        for row in load_sheet(args.listing, sheets["cm"]):
            sid = subject_key(row, sheets["cm"])
            name = val(row, sheets["cm"], "drug_name")
            if sid and sid not in skip_subjects and name and not is_research_drug(name, patterns):
                grouped["cm"][sid].append(row)
                centers.setdefault(sid, center_value(row, sheets["cm"]))
                statuses.setdefault(sid, subject_status(row, sheets["cm"]))
    if modules.get("pk") and sheets.get("pk"):
        pk_path = args.pk_listing or args.listing
        for row in load_sheet(pk_path, sheets["pk"]):
            sid = subject_key(row, sheets["pk"])
            if sid and sid not in skip_subjects:
                grouped["pk"][sid].append(row)
    if modules.get("findings") and sheets.get("findings") and args.finding_listing:
        for row in load_sheet(args.finding_listing, sheets["findings"]):
            sid = subject_key(row, sheets["findings"])
            if sid and sid not in skip_subjects:
                grouped["findings"][sid].append(row)

    subjects = sorted(set(visits) | set(grouped["ae"]) | set(grouped["mh"]) | set(grouped["cm"]) | set(grouped["pk"]) | set(grouped["findings"]))
    if population == "randomized":
        subjects = [sid for sid in subjects if group_info.get(sid, {}).get("is_randomized")]
    stats = {"subjects": len(subjects), "centers": len(set(centers.get(s, "未识别中心") for s in subjects)), "cm_prior": 0, "cm_concomitant": 0, "cm_reason_needs_confirmation": 0, "missing_windows": 0, "population": population}

    cards = []
    for sid in subjects:
        center = centers.get(sid, "未识别中心")
        display_sid = display_subject(sid, group_info, statuses.get(sid, ""))
        win_start, win_end = window_from_visits(sid, visits, sheets["visits"], config.get("study_window", {}), visit_order)
        hist_start, hist_end = window_from_visits(sid, visits, sheets["visits"], config.get("history_window", config.get("study_window", {})), visit_order)
        if not win_start or not win_end:
            stats["missing_windows"] += 1
        cm_prior, cm_con = [], []
        cm_cfg = sheets.get("cm", {})
        for row in grouped["cm"].get(sid, []):
            if cm_reason_needs_confirmation(row, cm_cfg):
                stats["cm_reason_needs_confirmation"] += 1
            if overlaps_window(row, cm_cfg, col(cm_cfg, "start"), col(cm_cfg, "end"), col(cm_cfg, "ongoing"), win_start, win_end):
                cm_con.append(row)
            else:
                cm_prior.append(row)
        stats["cm_prior"] += len(cm_prior)
        stats["cm_concomitant"] += len(cm_con)

        ae_rows = grouped["ae"].get(sid, [])
        mh_rows = grouped["mh"].get(sid, [])
        svg = render_svg(sid, visits.get(sid, []), ae_rows, mh_rows, cm_con, sheets, win_start, win_end, visit_order)

        sections = []
        if modules.get("ae", True):
            ae_cfg = sheets.get("ae", {})
            sections.append('<div class="box"><h3>AE</h3>' + render_list([
                li(f"<strong>{esc(val(r, ae_cfg, 'term'))}</strong>｜{esc(val(r, ae_cfg, 'start'))} ~ {esc(val(r, ae_cfg, 'end'))}｜{esc(val(r, ae_cfg, 'grade'))}｜关系 {esc(val(r, ae_cfg, 'relationship'))}｜SAE {esc(val(r, ae_cfg, 'sae'))}｜转归 {esc(val(r, ae_cfg, 'outcome'))}")
                for r in ae_rows
            ], "未检出 AE") + "</div>")
        if modules.get("mh", True):
            mh_cfg = sheets.get("mh", {})
            current_mh, past_mh = [], []
            for r in mh_rows:
                is_current = overlaps_window(r, mh_cfg, col(mh_cfg, "start"), col(mh_cfg, "end"), col(mh_cfg, "ongoing"), hist_start, hist_end)
                (current_mh if is_current else past_mh).append(r)
            sections.append(
                '<div class="box"><h3>现病史</h3>' + render_list([
                    li(f"<strong>{esc(val(r, mh_cfg, 'term'))}</strong>｜{esc(val(r, mh_cfg, 'start'))} ~ {esc(val(r, mh_cfg, 'end'))}")
                    for r in current_mh
                ], "无现病史")
                + '<h3>既往史</h3>' + render_list([
                    li(f"<strong>{esc(val(r, mh_cfg, 'term'))}</strong>｜{esc(val(r, mh_cfg, 'start'))} ~ {esc(val(r, mh_cfg, 'end'))}")
                    for r in past_mh
                ], "无既往史") + "</div>"
            )
        if modules.get("cm", True):
            sections.append('<div class="box"><h3>既往用药</h3>' + render_list([
                li(f"<strong>{esc(val(r, cm_cfg, 'drug_name'))}</strong>｜{esc(val(r, cm_cfg, 'start'))} ~ {esc(val(r, cm_cfg, 'end'))}｜{esc(val(r, cm_cfg, 'frequency'))}｜{esc(val(r, cm_cfg, 'route'))}｜{esc(cm_reason(r, cm_cfg))}")
                for r in cm_prior
            ], "无记录") + "</div>")
            sections.append('<div class="box"><h3>合并用药</h3>' + render_list([
                li(f"<strong>{esc(val(r, cm_cfg, 'drug_name'))}</strong>｜{esc(val(r, cm_cfg, 'start'))} ~ {esc(val(r, cm_cfg, 'end'))}｜{esc(val(r, cm_cfg, 'frequency'))}｜{esc(val(r, cm_cfg, 'route'))}｜{esc(cm_reason(r, cm_cfg))}")
                for r in cm_con
            ], "无记录") + "</div>")
        if modules.get("pk") and sheets.get("pk"):
            pk_cfg = sheets["pk"]
            sections.append('<div class="box"><h3>PK</h3>' + render_list([
                li(f"<strong>{esc(val(r, pk_cfg, 'visit'))}</strong>｜{esc(val(r, pk_cfg, 'result'))}｜LLOQ {esc(val(r, pk_cfg, 'lloq'))}｜{esc(val(r, pk_cfg, 'note'))}")
                for r in grouped["pk"].get(sid, [])
            ], "无 PK 记录") + "</div>")
        if modules.get("findings") and sheets.get("findings"):
            f_cfg = sheets["findings"]
            sections.append('<details class="detail-box" open><summary>Finding</summary>' + render_list([
                li(f"<strong>{esc(val(r, f_cfg, 'category'))}</strong>｜{esc(val(r, f_cfg, 'visit'))}｜{esc(val(r, f_cfg, 'severity'))}｜{esc(val(r, f_cfg, 'description'))}")
                for r in grouped["findings"].get(sid, [])
            ], "无 finding 记录") + "</details>")

        subtitle = f"{esc(center)}｜研究窗口：{esc(win_start)} ~ {esc(win_end)}"
        cards.append(f'''
<section class="subject-card" data-subject="{esc(sid)}" data-center="{esc(center)}">
  <div class="subject-top"><h2>{esc(display_sid)}</h2><div class="subject-sub">{subtitle}</div></div>
  {svg}
  <div class="cols">{"".join(sections)}</div>
</section>''')

    center_options = "".join(f'<option value="{esc(c)}">{esc(c)}</option>' for c in sorted(set(centers.get(s, "未识别中心") for s in subjects)))
    subject_options = "".join(f'<option value="{esc(s)}">{esc(display_subject(s, group_info, statuses.get(s, "")))}</option>' for s in subjects)
    html_text = f'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>{esc(config.get("title", "受试者时间线"))}</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,'PingFang SC','Microsoft YaHei',sans-serif;margin:0;background:#faf8f3;color:#222}}
header{{position:sticky;top:0;background:#fff;border-bottom:1px solid #eee7db;padding:10px 16px;z-index:10}}
.header-row{{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}}.toolbar{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
input,select{{padding:8px 10px;border:1px solid #d6d1c4;border-radius:10px;background:#fff;font-size:13px}}select{{min-width:170px}}
.wrap{{padding:20px;display:grid;gap:18px}}.subject-card{{background:#fff;border:1px solid #ece7dc;border-radius:18px;padding:18px;box-shadow:0 6px 18px rgba(0,0,0,.04)}}
.subject-top{{display:flex;justify-content:space-between;gap:16px;align-items:flex-start}}.subject-top h2{{margin:0;font-size:22px}}.subject-sub{{color:#666;font-size:13px;line-height:1.7}}
.svgline{{width:100%;height:auto;display:block;background:#fff;border:1px solid #eee7db;border-radius:12px;padding:8px;margin:12px 0}}.cols{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}
.box,.detail-box{{border:1px solid #f0e6d2;border-radius:12px;padding:12px;background:#fffdf8}}.box h3{{margin:0 0 8px;color:#9a5a00}}ul{{margin:0;padding-left:18px;font-size:12px;line-height:1.6}}
summary{{cursor:pointer;font-weight:700}}@media (max-width:960px){{.cols{{grid-template-columns:1fr}}.subject-top{{display:block}}}}
</style></head><body>
<header><div class="header-row"><strong>{esc(config.get("title", "受试者时间线"))}</strong><div class="toolbar">
<input id="q" placeholder="搜索"><select id="centerFilter"><option value="">全部中心</option>{center_options}</select><select id="subjectFilter"><option value="">全部受试者</option>{subject_options}</select>
</div></div></header><main class="wrap">
{"".join(cards)}
</main><script>
function applyFilters(){{const q=document.getElementById('q').value.trim().toLowerCase();const c=document.getElementById('centerFilter').value;const s=document.getElementById('subjectFilter').value;document.querySelectorAll('.subject-card').forEach(card=>{{const okQ=!q||card.innerText.toLowerCase().includes(q);const okC=!c||card.dataset.center===c;const okS=!s||card.dataset.subject===s;card.style.display=(okQ&&okC&&okS)?'block':'none';}})}}
['q','centerFilter','subjectFilter'].forEach(id=>document.getElementById(id).addEventListener('input',applyFilters));
</script></body></html>'''
    Path(args.output).write_text(html_text, encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
